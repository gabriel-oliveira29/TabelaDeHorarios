import openpyxl
import PySimpleGUI as sg
from datetime import datetime, timedelta
import locale
import os  # Importando a biblioteca os

# Definindo a localidade para português do Brasil
locale.setlocale(locale.LC_TIME, 'pt_BR.UTF-8')

# Função para formatar a data dinamicamente (DD/MM/AAAA) enquanto o usuário digita
def formatar_data_dinamica(valor):
    valor = valor.replace('/', '')  # Remove qualquer '/' existente
    if len(valor) > 2:
        valor = valor[:2] + '/' + valor[2:]  # Adiciona '/' após o dia
    if len(valor) > 5:
        valor = valor[:5] + '/' + valor[5:]  # Adiciona '/' após o mês
    return valor

# Função para formatar horas dinamicamente (HH:MM)
def formatar_hora_dinamica(valor):
    valor = valor.replace(':', '')  # Remove qualquer ':' existente
    if len(valor) > 2:
        valor = valor[:2] + ':' + valor[2:]  # Adiciona ':' após os dois primeiros dígitos
    return valor

# Função para formatar a data
def data_formatada(data):
    if isinstance(data, datetime):
        dia = data.strftime('%d')
        mes = data.strftime('%m')
        return f"{dia}/{mes}"
    return "Data inválida"

# Função para obter o dia da semana em português
def obter_dia_da_semana(data):
    dias_da_semana = [
        'segunda-feira', 'terça-feira', 'quarta-feira',
        'quinta-feira', 'sexta-feira', 'sábado', 'domingo'
    ]
    return dias_da_semana[data.weekday()]

# Interface gráfica
sg.theme('reddit')
layout_funcionario = [
    [sg.Text('Selecione o setor:')],
    [sg.InputOptionMenu(('Patio', 'Escritório', 'Estagiário - 6HR', 'Estagiário - 3h30'), key='setor')],
    [sg.Text('Nome do Funcionário')],
    [sg.Input(key='NomeFuncionario')],
    [sg.Text('Data de Início (DD/MM/AAAA)')],
    [sg.Input(key='datainicio', enable_events=True)],  # Adicionando eventos para a data de início
    [sg.Text('Mês')],
    [sg.Input(key='Mes')],
    [sg.Text('Ano')],
    [sg.Input(key='Ano')],
    [sg.Button('Continuar', bind_return_key=True), sg.Button('Sair')]
]

layout_linha = [
    [sg.Text('Data: '), sg.Text('', key='data'), sg.Text('', key='diasemana')],
    [sg.Text('Hora Entrada')],
    [sg.Input(key='hora_entrada', enable_events=True)],
    [sg.Text('Saída para Almoço')],
    [sg.Input(key='saida_almoco', enable_events=True)],
    [sg.Text('Retorno do Almoço')],
    [sg.Input(key='retorno_almoco', enable_events=True)],
    [sg.Text('Saída')],
    [sg.Input(key='saida', enable_events=True)],
    [sg.Button('Salvar', bind_return_key=True), sg.Button('Cancelar')]
]

# Variável para controlar a linha atual
linha_atual = 4
            
# Janela para selecionar o funcionário
janela_funcionario = sg.Window('Seleção de Funcionário', layout=layout_funcionario)

while True:
    event, value = janela_funcionario.read()

    if event == sg.WIN_CLOSED or event == 'Sair':
        break

    # Atualizando o campo da data dinamicamente enquanto o usuário digita
    if event == 'datainicio':
        janela_funcionario['datainicio'].update(formatar_data_dinamica(value['datainicio']))

    if event == 'Continuar':
        setor = value['setor']
        NomeFuncionario = value['NomeFuncionario']

        # Selecionando a planilha correta com base no setor
        if setor == 'Patio':
            workbook = openpyxl.load_workbook('HorariosPatio.xlsx')
            planilha = workbook['Teste']
        elif setor == 'Escritório':
            workbook = openpyxl.load_workbook('HorariosEscritorio.xlsx')
            planilha = workbook['Teste']
        elif setor == 'Estagiário - 6HR':
            workbook = openpyxl.load_workbook('HorariosEstagiario6hr.xlsx')
            planilha = workbook['Teste']
        elif setor == 'Estagiário - 3h30':
            workbook = openpyxl.load_workbook('HorariosEstagiario3h30.xlsx')
            planilha = workbook['Teste']

        # Lógica para processar a planilha carregada
        planilha['A1'].value = NomeFuncionario

        # Criando uma pasta para armazenar os arquivos na área de trabalho, se não existir
        desktop = os.path.join(os.path.expanduser("~"), "Desktop")  # Caminho da área de trabalho
        pasta_arquivos = os.path.join(desktop, 'ArquivosFuncionarios')  # Nome da pasta na área de trabalho
        if not os.path.exists(pasta_arquivos):
            os.makedirs(pasta_arquivos)

        mes = value['Mes']
        planilha['G1'].value = mes

        ano = value['Ano']
        planilha['J1'].value = ano

        data_inicio = value['datainicio'].replace('/', '')  # Removendo as barras para processar a data
        data_inicio_dt = datetime.strptime(data_inicio, "%d%m%Y")

        # Atualiza a data inicial na planilha
        for row in planilha.iter_rows(min_row=4, max_row=34, min_col=2, max_col=2):
            for cell in row:
                cell.value = data_inicio_dt
                data_inicio_dt += timedelta(days=1)  # Incrementando um dia
                
        # Salvando o arquivo na nova pasta
        workbook.save(os.path.join(pasta_arquivos, f'{NomeFuncionario}.xlsx'))
        # Fechando a primeira janela
        janela_funcionario.close()

        # Inicializando a janela de entrada por linha
        janela_linha = sg.Window('Preencher Horários', layout=layout_linha)

        # Reinicializando a data para a data inicial lida da planilha
        data = planilha['B4'].value

        while linha_atual <= 34:
            event, value = janela_linha.read()

            if event == sg.WIN_CLOSED or event == 'Cancelar':
                break

            # Atualizando os campos de hora dinamicamente enquanto o usuário digita
            if event == 'hora_entrada':
                janela_linha['hora_entrada'].update(formatar_hora_dinamica(value['hora_entrada']))

            if event == 'saida_almoco':
                janela_linha['saida_almoco'].update(formatar_hora_dinamica(value['saida_almoco']))

            if event == 'retorno_almoco':
                janela_linha['retorno_almoco'].update(formatar_hora_dinamica(value['retorno_almoco']))

            if event == 'saida':
                janela_linha['saida'].update(formatar_hora_dinamica(value['saida']))

            if event == 'Salvar':
                # Validação das horas
                HoraEntrada = value['hora_entrada'].replace(":", "")
                SaidaAlmoco = value['saida_almoco'].replace(":", "")
                RetornoAlmoco = value['retorno_almoco'].replace(":", "")
                Saida = value['saida'].replace(":", "")

                if len(HoraEntrada) != 4 or not HoraEntrada.isdigit():
                    sg.popup_error("Hora de entrada inválida. Use o formato HHMM.")
                    continue
                if len(SaidaAlmoco) != 4 or not SaidaAlmoco.isdigit():
                    sg.popup_error("Hora de saída para almoço inválida. Use o formato HHMM.")
                    continue
                if len(RetornoAlmoco) != 4 or not RetornoAlmoco.isdigit():
                    sg.popup_error("Hora de retorno do almoço inválida. Use o formato HHMM.")
                    continue
                if len(Saida) != 4 or not Saida.isdigit():
                    sg.popup_error("Hora de saída inválida. Use o formato HHMM.")
                    continue

                # Verificando se o Retorno do Almoço é menor que a Saída para Almoço
                if RetornoAlmoco < SaidaAlmoco:
                    sg.popup_error("O horário de Retorno do Almoço não pode ser mais cedo que saida almoço")
                    continue

                # Atualizando a planilha
                row = planilha[linha_atual]
                row[2].value = HoraEntrada[:2] + ":" + HoraEntrada[2:]
                row[3].value = SaidaAlmoco[:2] + ":" + SaidaAlmoco[2:]
                row[4].value = RetornoAlmoco[:2] + ":" + RetornoAlmoco[2:]
                row[5].value = Saida[:2] + ":" + Saida[2:]

                # Incrementando a data para a próxima linha
                data += timedelta(days=1)

                # Verificando o tipo do valor lido e formatando
                if isinstance(data, datetime):
                    data_F = data_formatada(data)
                    dia_da_semana = obter_dia_da_semana(data)  # Obter o dia da semana

                # Atualizando a linha na interface para refletir corretamente
                janela_linha['data'].update(f"Data: {data_F}")
                janela_linha['diasemana'].update(f"Dia da Semana: {dia_da_semana}")

                # Salvando o arquivo Excel
                workbook.save(os.path.join(pasta_arquivos, f'{NomeFuncionario}.xlsx'))

                # Atualiza para a próxima linha
                linha_atual += 1

                # Limpar os inputs para a próxima linha
                janela_linha['hora_entrada'].update('')
                janela_linha['saida_almoco'].update('')
                janela_linha['retorno_almoco'].update('')
                janela_linha['saida'].update('')

                # Mover o foco de volta para o campo de Hora de Entrada
                janela_linha['hora_entrada'].set_focus()

        janela_linha.close()  # Fecha a janela de preenchimento

janela_funcionario.close()
